"""
OpenInsight — Router Monitoring Tool
Polls a Linksys/OpenWRT router via SSH every 60 seconds,
captures system health metrics, and logs them to an Excel
file with a live dashboard chart.

Security Note: Credentials are hardcoded per project spec.
For production use, store credentials in a .env file and
load them with python-dotenv.
"""

import os
import time
from datetime import datetime

import paramiko
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Router Configuration ──────────────────────────────────────
ROUTER_IP = "192.168.150.1"
ROUTER_USER = "root"
ROUTER_PASS = "Sherland130l!"
EXCEL_FILE = "router_report.xlsx"
POLL_INTERVAL = 60  # seconds


def connect_ssh():
    """Establish an SSH connection to the router.

    Returns the SSHClient on success, or None on failure.
    Prints [+] on success, [!] with reason on failure.
    """
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    try:
        client.connect(
            hostname=ROUTER_IP,
            username=ROUTER_USER,
            password=ROUTER_PASS,
            timeout=10,
        )
        print(f"[+] SSH connection established to {ROUTER_IP}")
        return client
    except paramiko.AuthenticationException:
        print(f"[!] Auth Error: Invalid credentials for {ROUTER_USER}@{ROUTER_IP}")
    except (OSError, paramiko.ssh_exception.NoValidConnectionsError) as e:
        print(f"[!] Network Unreachable: Cannot reach {ROUTER_IP} — {e}")
    except paramiko.SSHException as e:
        print(f"[!] SSH Error: {e}")

    return None


def _exec_command(ssh_client, command, label):
    """Run a single command over SSH and return its stdout.

    Prints [>] status per command.  Returns None on timeout/error.
    """
    try:
        stdin, stdout, stderr = ssh_client.exec_command(command, timeout=10)
        output = stdout.read().decode().strip()
        print(f"  [>] {label}: OK")
        return output
    except Exception as e:
        print(f"  [>] {label}: TIMEOUT/ERROR — {e}")
        return None


def harvest_data(ssh_client):
    """Execute monitoring commands and parse their output.

    Returns a dict with timestamp, cpu_load, process_count,
    and top_process, or None if critical data is missing.
    """
    print("[>] Harvesting data …")

    # 1. CPU load (1-min average from /proc/loadavg)
    loadavg_raw = _exec_command(ssh_client, "cat /proc/loadavg", "CPU Load")
    cpu_load = None
    if loadavg_raw:
        try:
            cpu_load = float(loadavg_raw.split()[0])
        except (IndexError, ValueError):
            print("  [!] Could not parse /proc/loadavg")

    # 2. Active process count
    ps_raw = _exec_command(ssh_client, "ps | wc -l", "Process Count")
    process_count = None
    if ps_raw:
        try:
            process_count = int(ps_raw.strip())
        except ValueError:
            print("  [!] Could not parse process count")

    # 3. Top 10 processes by CPU usage from top output
    # BusyBox top format: PID PPID USER STAT VSZ %VSZ CPU %CPU COMMAND...
    #                      0    1    2    3   4   5    6   7    8+
    # Lines 0-3 are headers; process list starts at line 4.
    # COMMAND may contain spaces, so join parts[8:].
    top_raw = _exec_command(ssh_client, "top -bn1 | sed -n '5,14p'", "Top 10 Processes")
    top_processes = []
    if top_raw:
        for line in top_raw.splitlines():
            parts = line.split()
            if len(parts) >= 9:
                try:
                    cpu_pct = float(parts[7].strip("%"))
                except ValueError:
                    cpu_pct = 0.0
                try:
                    mem_pct = float(parts[5].strip("%"))
                except ValueError:
                    mem_pct = 0.0
                proc_name = " ".join(parts[8:])
                top_processes.append({
                    "name": proc_name,
                    "cpu_pct": cpu_pct,
                    "mem_pct": mem_pct,
                })

    # Pad to exactly 10 entries so Excel columns stay consistent
    while len(top_processes) < 10:
        top_processes.append({"name": None, "cpu_pct": None, "mem_pct": None})

    data = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "cpu_load": cpu_load,
        "process_count": process_count,
        "top_processes": top_processes[:10],
    }
    return data


def _ensure_sheet(wb, name, headers):
    """Return a sheet by *name*, creating it with *headers* if absent."""
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(headers)
    else:
        ws = wb[name]
    return ws


def update_excel(data_dict):
    """Append data to Logs and Top Processes sheets in router_report.xlsx.

    Creates the workbook and sheets if they don't exist.
    Returns the workbook on success so update_dashboard can use it,
    or None if the file is locked.
    """
    logs_headers = ["Timestamp", "CPU Load (1-min)", "Process Count"]
    tp_headers = ["Timestamp"]
    for i in range(1, 11):
        tp_headers += [f"#{i} Name", f"#{i} CPU%", f"#{i} MEM%"]

    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError:
            print("!! FILE LOCKED: Please close Excel to save data !!")
            return None
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"
        ws.append(logs_headers)
        print("[+] Created new workbook with Logs sheet")

    # ── Logs sheet (overall health) ──
    logs_ws = _ensure_sheet(wb, "Logs", logs_headers)
    logs_ws.append([
        data_dict["timestamp"],
        data_dict["cpu_load"],
        data_dict["process_count"],
    ])
    logs_row = logs_ws.max_row

    # ── Top Processes sheet (per-process CPU% & MEM%) ──
    tp_ws = _ensure_sheet(wb, "Top Processes", tp_headers)
    tp_row = [data_dict["timestamp"]]
    for proc in data_dict["top_processes"]:
        tp_row += [proc["name"], proc["cpu_pct"], proc["mem_pct"]]
    tp_ws.append(tp_row)

    try:
        wb.save(EXCEL_FILE)
        print(f"[+] Data saved — Logs row {logs_row}, "
              f"Top Processes row {tp_ws.max_row}")
    except PermissionError:
        print("!! FILE LOCKED: Please close Excel to save data !!")
        return None

    return wb


def update_dashboard(workbook):
    """Maintain a Dashboard sheet with four charts:
    1. CPU Load (from Logs)
    2. Process Count (from Logs)
    3. Top 10 CPU% (from Top Processes)
    4. Top 10 MEM% (from Top Processes)
    """
    if "Logs" not in workbook.sheetnames:
        return

    logs_ws = workbook["Logs"]
    logs_max = logs_ws.max_row

    if logs_max < 2:
        print("[!] Not enough data for dashboard charts yet")
        return

    # Create or clear the Dashboard sheet
    if "Dashboard" in workbook.sheetnames:
        workbook.remove(workbook["Dashboard"])
    dash_ws = workbook.create_sheet("Dashboard")

    # ── Chart 1: CPU Load (Logs col B) ──
    categories = Reference(logs_ws, min_col=1, min_row=2, max_row=logs_max)

    cpu_chart = LineChart()
    cpu_chart.title = "CPU Load (1-min avg)"
    cpu_chart.style = 10
    cpu_chart.y_axis.title = "Load"
    cpu_chart.x_axis.title = "Time"
    cpu_chart.y_axis.delete = False
    cpu_chart.x_axis.delete = False
    cpu_chart.x_axis.tickLblPos = "low"
    cpu_chart.y_axis.tickLblPos = "low"
    cpu_chart.width = 28
    cpu_chart.height = 14
    cpu_chart.add_data(Reference(logs_ws, min_col=2, min_row=1,
                                 max_row=logs_max), titles_from_data=True)
    cpu_chart.set_categories(categories)
    dash_ws.add_chart(cpu_chart, "A1")

    # ── Chart 2: Process Count (Logs col C) ──
    proc_chart = LineChart()
    proc_chart.title = "Active Process Count"
    proc_chart.style = 10
    proc_chart.y_axis.title = "Count"
    proc_chart.x_axis.title = "Time"
    proc_chart.y_axis.delete = False
    proc_chart.x_axis.delete = False
    proc_chart.x_axis.tickLblPos = "low"
    proc_chart.y_axis.tickLblPos = "low"
    proc_chart.width = 28
    proc_chart.height = 14
    proc_chart.add_data(Reference(logs_ws, min_col=3, min_row=1,
                                  max_row=logs_max), titles_from_data=True)
    proc_chart.set_categories(categories)
    dash_ws.add_chart(proc_chart, "A35")

    # ── Per-process charts (Top Processes sheet) ──
    if "Top Processes" not in workbook.sheetnames:
        return
    tp_ws = workbook["Top Processes"]
    tp_max = tp_ws.max_row
    if tp_max < 2:
        return

    tp_categories = Reference(tp_ws, min_col=1, min_row=2, max_row=tp_max)

    # Resolve process names from the latest row — use basename only
    proc_labels = []
    for i in range(10):
        name_col = 2 + i * 3  # Name columns: 2, 5, 8, …, 29
        raw = tp_ws.cell(row=tp_max, column=name_col).value
        if raw:
            first_token = raw.split()[0]          # drop arguments
            short = os.path.basename(first_token)  # drop path
            # strip wrapping like {name} or [name]
            short = short.strip("{}[]")
            proc_labels.append(short)
        else:
            proc_labels.append(f"#{i + 1}")

    # ── Chart 3: Top 10 CPU% (cols 3,6,9,...,30) ──
    cpu_pct_chart = LineChart()
    cpu_pct_chart.title = "Top 10 Processes — CPU%"
    cpu_pct_chart.style = 10
    cpu_pct_chart.y_axis.title = "CPU %"
    cpu_pct_chart.x_axis.title = "Time"
    cpu_pct_chart.y_axis.delete = False
    cpu_pct_chart.x_axis.delete = False
    cpu_pct_chart.x_axis.tickLblPos = "low"
    cpu_pct_chart.y_axis.tickLblPos = "low"
    cpu_pct_chart.width = 28
    cpu_pct_chart.height = 14
    for i in range(10):
        col = 3 + i * 3  # columns 3, 6, 9, …, 30
        data_ref = Reference(tp_ws, min_col=col, min_row=1, max_row=tp_max)
        cpu_pct_chart.add_data(data_ref, titles_from_data=True)
        cpu_pct_chart.series[-1].tx = SeriesLabel(v=proc_labels[i])
    cpu_pct_chart.set_categories(tp_categories)
    dash_ws.add_chart(cpu_pct_chart, "A69")

    # ── Chart 4: Top 10 MEM% (cols 4,7,10,...,31) ──
    mem_pct_chart = LineChart()
    mem_pct_chart.title = "Top 10 Processes — MEM%"
    mem_pct_chart.style = 10
    mem_pct_chart.y_axis.title = "MEM %"
    mem_pct_chart.x_axis.title = "Time"
    mem_pct_chart.y_axis.delete = False
    mem_pct_chart.x_axis.delete = False
    mem_pct_chart.x_axis.tickLblPos = "low"
    mem_pct_chart.y_axis.tickLblPos = "low"
    mem_pct_chart.width = 28
    mem_pct_chart.height = 14
    for i in range(10):
        col = 4 + i * 3  # columns 4, 7, 10, …, 31
        data_ref = Reference(tp_ws, min_col=col, min_row=1, max_row=tp_max)
        mem_pct_chart.add_data(data_ref, titles_from_data=True)
        mem_pct_chart.series[-1].tx = SeriesLabel(v=proc_labels[i])
    mem_pct_chart.set_categories(tp_categories)
    dash_ws.add_chart(mem_pct_chart, "A103")

    try:
        workbook.save(EXCEL_FILE)
        print("[+] Dashboard refreshed (4 charts)")
    except PermissionError:
        print("!! FILE LOCKED: Could not update dashboard !!")


def main():
    """Main polling loop — runs every 60 seconds until Ctrl+C."""
    print("=" * 50)
    print("  OpenInsight — Router Monitor")
    print(f"  Target: {ROUTER_IP}")
    print(f"  Interval: {POLL_INTERVAL}s")
    print("=" * 50)

    try:
        while True:
            print(f"\n--- Cycle @ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---")

            ssh = connect_ssh()
            if ssh is None:
                print("[!] Skipping cycle — no SSH connection")
                time.sleep(POLL_INTERVAL)
                continue

            try:
                data = harvest_data(ssh)
            finally:
                ssh.close()
                print("[+] SSH session closed")

            if data is None:
                print("[!] Skipping cycle — harvest failed")
                time.sleep(POLL_INTERVAL)
                continue

            wb = update_excel(data)
            if wb is not None:
                update_dashboard(wb)

            print(f"[+] Sleeping {POLL_INTERVAL}s …")
            time.sleep(POLL_INTERVAL)

    except KeyboardInterrupt:
        print("\n[+] Monitoring stopped by user. Goodbye!")


if __name__ == "__main__":
    main()
